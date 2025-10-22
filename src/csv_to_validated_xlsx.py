#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse, csv, os
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Required input CSV columns
REQ_COLS = ["qa_id", "question", "expected_answer", "source_text", "target_text"]

# CSV label columns (keep whatever values are there)
ANN_COLS_OLD = [
    "1_question_valid",
    "2_needed",
    "3_answer_correct",
    "4_final_decision",
    "5_comment",
]

# Output Excel headers (long form, matching the HTML)
Q1 = "1) Is the question valid and well-formed?"
Q2 = "2) Which text(s) are needed to answer?"
Q3 = "3) Is the provided answer correct and complete?"
Q4 = "4) Final Decision"
Q5 = "5) Optional comment"

FINAL_HEADERS = [
    "qa_id",
    "question",
    "source_text",
    "target_text",
    "expected_answer",
    Q1, Q2, Q3, Q4, Q5,
]

# Column widths (tweak if you like)
COL_WIDTHS = {
    "qa_id": 26,
    "question": 46,
    "source_text": 54,
    "target_text": 54,
    "expected_answer": 46,
    Q1: 42, Q2: 44, Q3: 48, Q4: 20, Q5: 36,
}

INSTR_TEXT = """Regulatory Question & Answer Annotation — Quick Guide

Goal
Evaluate if the Proposed Answer correctly and completely answers the Question, using ONLY Rule 1 and Rule 2. Do not use external knowledge.

Columns
- question / source_text / target_text / expected_answer: Read carefully.
- 1) Is the question valid and well-formed?
- 2) Which text(s) are needed to answer?
- 3) Is the provided answer correct and complete?
- 4) Final Decision
- 5) Optional comment

Tips
- Keep comments brief (1–2 clauses).
- Cite which rule(s) inform your decision if helpful (e.g., “Both: R2 lists records; R1 links requestability”).
"""

def read_csv(path: str) -> List[Dict[str, str]]:
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = [dict(r) for r in reader]
    return rows

def build_workbook(rows: List[Dict[str, str]]) -> Workbook:
    wb = Workbook()

    # Instructions first (sheet 1)
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info["A1"] = INSTR_TEXT
    ws_info["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_info.column_dimensions["A"].width = 120
    ws_info.row_dimensions[1].height = 520

    # Data sheet (sheet 2)
    ws = wb.create_sheet("Annotations")

    # Header row
    ws.append(FINAL_HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="263371")
    for c in range(1, len(FINAL_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill

    # Wrap text, freeze header, enable filter
    wrap = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FINAL_HEADERS))}1"

    # Ensure annotation columns exist even if empty
    for r in rows:
        for k in ANN_COLS_OLD:
            r.setdefault(k, "")

    # Write rows (preserve whatever label values exist)
    for r in rows:
        ws.append([
            r.get("qa_id", ""),
            r.get("question", ""),
            r.get("source_text", ""),
            r.get("target_text", ""),
            r.get("expected_answer", ""),
            r.get("1_question_valid", ""),
            r.get("2_needed", ""),
            r.get("3_answer_correct", ""),
            r.get("4_final_decision", ""),
            r.get("5_comment", ""),
        ])

    # Column widths + wrap
    for name, width in COL_WIDTHS.items():
        j = FINAL_HEADERS.index(name) + 1
        ws.column_dimensions[get_column_letter(j)].width = width
    for row in ws.iter_rows(min_row=2, max_col=len(FINAL_HEADERS)):
        for cell in row:
            cell.alignment = wrap

    return wb

def main():
    ap = argparse.ArgumentParser(description="Convert ObliQA CSV to a clean XLSX (no dropdowns).")
    ap.add_argument("input_csv", help="Path to input CSV")
    ap.add_argument("output_xlsx", help="Path to output XLSX")
    args = ap.parse_args()

    rows = read_csv(args.input_csv)
    if not rows:
        raise SystemExit("Input CSV has no rows.")
    # sanity check required columns
    missing = [c for c in REQ_COLS if c not in rows[0].keys()]
    if missing:
        raise SystemExit(f"Missing required columns in CSV: {missing}")

    wb = build_workbook(rows)
    wb.save(args.output_xlsx)
    print(f"✅ Wrote {os.path.basename(args.output_xlsx)} (no dropdowns).")

if __name__ == "__main__":
    main()
